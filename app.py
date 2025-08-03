import streamlit as st
import pandas as pd
import re
import csv
import time
import os
import glob
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import NamedStyle


# ============================================================================
# FUNCIÓN DE REDONDEO AGRESIVO
# ============================================================================


def redondear_agresivo(valor):
    """Redondea agresivamente a 2 decimales usando Decimal para mayor precisión"""
    if pd.isna(valor) or valor == 0 or valor == "":
        return valor
    try:
        # Convertir a Decimal para mayor precisión y luego redondear
        decimal_val = Decimal(str(float(valor))).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
        return float(decimal_val)
    except (ValueError, TypeError, InvalidOperation):
        return valor


# ============================================================================
# FUNCIONES DE LECTURA Y LIMPIEZA DE ARCHIVOS
# ============================================================================


def leer_archivo(file_path):
    """Lee el archivo con manejo de codificación UTF-8 o Latin-1"""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return f.readlines()
    except UnicodeDecodeError:
        with open(file_path, "r", encoding="latin-1") as f:
            return f.readlines()


def procesar_encabezado(lines):
    """Procesa y extrae la información del encabezado del archivo"""
    try:
        encabezado = lines[1:7]
        encabezado_limpio = [line.replace("\n", "").strip() for line in encabezado]

        return {
            "RAZON SOCIAL": encabezado_limpio[0],
            "DIRECCION": encabezado_limpio[1],
            "CUIT": encabezado_limpio[2],
            "LIBRO": encabezado_limpio[3].split("  ")[-1],
            "PERIODO": encabezado_limpio[4].split("  ")[-1],
        }
    except Exception as e:
        st.error(f"Ocurrió un error al procesar el encabezado: {e}")
        return {}


def limpiar_lineas(lines):
    """Limpia las líneas del archivo eliminando caracteres de control y bloques no deseados"""
    cleaned_lines = []
    eliminar = False
    eliminar_desde_totales = False
    compras_o_ventas = ""

    for i, line in enumerate(lines[9:], start=2):
        # Detectar tipo de operación
        if "IVA VENTAS" in line:
            compras_o_ventas = "Ventas"
        elif "IVA COMPRAS" in line:
            compras_o_ventas = "Compras"

        # Detectar fin de datos
        if "TOTALES POR TASA" in line:
            eliminar_desde_totales = True
            continue

        if eliminar_desde_totales:
            break

        # Manejar bloques a eliminar
        if line.startswith("----"):
            eliminar = True
            continue

        if line.startswith("--"):
            eliminar = False
            continue

        # Procesar líneas válidas
        if not eliminar:
            cleaned_line = re.sub(r"\x1b[^m]*m", "", line)  # Elimina secuencias ANSI
            cleaned_line = re.sub(
                r"[\x00-\x1F\x7F]", "", cleaned_line
            )  # Eliminación de caracteres de control ASCII
            cleaned_lines.append(cleaned_line)

    return cleaned_lines, compras_o_ventas


def limpiar_lineas_adicional(cleaned_lines):
    """Segunda limpieza de líneas eliminando líneas con PPag"""
    doble_cleaned_lines = []

    for i, line in enumerate(cleaned_lines, start=-1):
        if "PPag." in line or len(line.strip()) < 35:
            if re.search(r"PPag\.\:\s*\d+\s*$", line):
                linea = re.sub(r"PPag\.\:\s*\d+\s*$", "", line)
                cleaned_lines.append(linea)
            else:
                break
        doble_cleaned_lines.append(line)

    return doble_cleaned_lines


# ============================================================================
# FUNCIONES DE PROCESAMIENTO DE MOVIMIENTOS
# ============================================================================


def procesar_movimientos(doble_cleaned_lines, compras_o_ventas):
    """Procesa las líneas limpias y extrae los movimientos"""
    movements = []
    temp_movement = {}

    for index, cleaned_line in enumerate(doble_cleaned_lines):
        # Procesar líneas continuas del mismo movimiento
        if "numero" in temp_movement and temp_movement["numero"] == cleaned_line[12:20]:
            procesar_linea_continuacion(cleaned_line, temp_movement, compras_o_ventas)
        else:
            # Procesar nueva línea de movimiento
            if cleaned_line[0:2] == "  ":
                procesar_linea_continuacion(
                    cleaned_line, temp_movement, compras_o_ventas
                )
                if index == len(doble_cleaned_lines) - 1:
                    movements.append(temp_movement)
            else:
                # Nueva entrada de movimiento
                movement = temp_movement.copy()
                movements.append(movement)
                temp_movement.clear()

                procesar_nueva_entrada(cleaned_line, temp_movement, compras_o_ventas)

                if index == len(doble_cleaned_lines) - 1:
                    movements.append(temp_movement)

    # Limpiar movimiento vacío inicial
    if not movements[0]:
        movements.pop(0)

    return movements


def procesar_linea_continuacion(cleaned_line, temp_movement, compras_o_ventas):
    """Procesa una línea que continúa un movimiento existente"""
    partes = re.split(r"\s{3,}", cleaned_line[70:])
    if len(partes) < 2:
        return

    tasa = partes[0]
    if tasa in [
        "Tasa 21%",
        "T.10.5%",
        "Tasa 27%",
        "C.F.21%",
        "C.F.10.5%",
        "Tasa 2.5%",
        "T.IMP 21%",
        "T.IMP 10%",
    ]:
        procesar_tasa_con_neto_iva(tasa, partes, temp_movement)
    elif tasa in ["R.Monot21", "R.Mont.10"]:
        procesar_tasa_monotributo(tasa, partes, temp_movement, compras_o_ventas)
    else:
        procesar_otra_tasa(tasa, partes, temp_movement)


def procesar_tasa_con_neto_iva(tasa, partes, temp_movement):
    """Procesa tasas que tienen neto e IVA separados"""
    if tasa + " Neto" in temp_movement:
        # Sumar valores existentes
        neto_anterior = float(temp_movement[tasa + " Neto"].replace(",", "."))
        iva_anterior = float(temp_movement[tasa + " IVA"].replace(",", "."))
        neto_nuevo = float(partes[1].replace(",", "."))
        iva_nuevo = float(partes[2].replace(",", "."))

        temp_movement[tasa + " Neto"] = round(neto_anterior + neto_nuevo, 2)
        temp_movement[tasa + " IVA"] = round(iva_anterior + iva_nuevo, 2)
    else:
        # Primer valor
        temp_movement[tasa + " Neto"] = partes[1]
        temp_movement[tasa + " IVA"] = partes[2]


def procesar_tasa_monotributo(tasa, partes, temp_movement, compras_o_ventas):
    """Procesa tasas de monotributo"""
    if compras_o_ventas == "Ventas":
        if tasa + " Neto" in temp_movement:
            neto_anterior = float(temp_movement[tasa + " Neto"].replace(",", "."))
            iva_anterior = float(temp_movement[tasa + " IVA"].replace(",", "."))
            neto_nuevo = float(partes[1].replace(",", "."))
            iva_nuevo = float(partes[2].replace(",", "."))

            temp_movement[tasa + " Neto"] = round(neto_anterior + neto_nuevo, 2)
            temp_movement[tasa + " IVA"] = round(iva_anterior + iva_nuevo, 2)
        else:
            temp_movement[tasa + " Neto"] = partes[1]
            temp_movement[tasa + " IVA"] = partes[2]
    else:
        temp_movement[tasa] = partes[1]


def procesar_otra_tasa(tasa, partes, temp_movement):
    """Procesa otras tasas que no tienen neto/IVA separados"""
    if tasa in temp_movement:
        if isinstance(temp_movement[tasa], float):
            numero_actual = float(partes[1].replace(",", "."))
            temp_movement[tasa] = round(temp_movement[tasa] + numero_actual, 2)
        else:
            numero_anterior = float(temp_movement[tasa].replace(",", "."))
            numero_actual = float(partes[1].replace(",", "."))
            temp_movement[tasa] = round(numero_anterior + numero_actual, 2)
    else:
        temp_movement[tasa] = partes[1]


def procesar_nueva_entrada(cleaned_line, temp_movement, compras_o_ventas):
    """Procesa una nueva entrada de movimiento"""
    partes = re.split(r"\s{3,}", cleaned_line[70:])
    if len(partes) < 2:
        return

    temp_movement.update(
        {
            "Fecha": cleaned_line[0:2],
            "Comprobante": cleaned_line[3:5],
            "PV": cleaned_line[6:11],
            "Nro": cleaned_line[12:20],
            "Letra": cleaned_line[20:21],
            "Razon Social": cleaned_line[22:44],
            "Condicion": cleaned_line[45:49],
            "CUIT": cleaned_line[50:63],
            "Concepto": cleaned_line[64:67],
            "Jurisdiccion": cleaned_line[68:69],
        }
    )

    # Procesar montos
    if len(partes) == 3:
        primer_monto = partes[1].split(" ")
        primer_monto = list(filter(None, primer_monto))
        segundo_monto = partes[1].split(" ")
        segundo_monto = list(filter(None, segundo_monto))
        partes = [partes[0]] + primer_monto + segundo_monto

    tasa = partes[0]
    if tasa in [
        "Tasa 21%",
        "T.10.5%",
        "Tasa 27%",
        "C.F.21%",
        "C.F.10.5%",
        "Tasa 2.5%",
        "T.IMP 21%",
        "T.IMP 10%",
    ]:
        temp_movement[tasa + " Neto"] = partes[1]
        temp_movement[tasa + " IVA"] = partes[2]
    elif tasa in ["R.Monot21", "R.Mont.10"]:
        if compras_o_ventas == "Ventas":
            temp_movement[tasa + " Neto"] = partes[1]
            temp_movement[tasa + " IVA"] = partes[2]
        else:
            temp_movement[tasa] = partes[1]
    else:
        temp_movement[tasa] = partes[1]


# ============================================================================
# FUNCIONES DE PROCESAMIENTO DE DATAFRAMES
# ============================================================================


def crear_dataframe_movimientos(movements):
    """Crea y procesa el DataFrame de movimientos"""
    df = pd.DataFrame(movements)
    df = df.fillna(0)

    # Reemplazar comas por puntos en columnas numéricas
    df.iloc[:, 11:] = df.iloc[:, 11:].replace(",", ".", regex=True)
    df.iloc[:, 11:] = df.iloc[:, 11:].apply(pd.to_numeric, errors="coerce").fillna(0)

    # Aplicar redondeo agresivo para eliminar errores de precisión desde el inicio
    for col in df.columns[11:]:
        df[col] = df[col].apply(redondear_agresivo)

    # Convertir notas de crédito a negativas
    columnas_a_convertir = df.columns[11:]
    df.loc[df["Comprobante"] == "NC", columnas_a_convertir] *= -1

    # Redondear nuevamente después de la multiplicación por -1
    for col in df.columns[11:]:
        df[col] = df[col].apply(redondear_agresivo)

    # Convertir tipos de datos
    df["PV"] = pd.to_numeric(df["PV"])
    df["Nro"] = pd.to_numeric(df["Nro"])
    df["Concepto"] = pd.to_numeric(df["Concepto"])

    return df


def combinar_movimientos_duplicados(df):
    """Combina movimientos que tienen la misma clave principal"""
    resultado = []
    fila_actual = df.iloc[0].copy()

    for i in range(1, len(df)):
        fila_siguiente = df.iloc[i]

        # Si la clave principal se repite, combinar valores
        if (
            fila_actual["Nro"] == fila_siguiente["Nro"]
            and fila_actual["PV"] == fila_siguiente["PV"]
            and fila_actual["Razon Social"] == fila_siguiente["Razon Social"]
        ):

            for col in df.columns[11:]:  # Sumar solo las columnas numéricas
                fila_actual[col] = redondear_agresivo(
                    fila_actual[col] + fila_siguiente[col]
                )
        else:
            resultado.append(fila_actual)
            fila_actual = fila_siguiente.copy()

    resultado.append(fila_actual)
    return pd.DataFrame(resultado)


def agregar_totales_movimientos(df_final):
    """Agrega fila de totales al DataFrame de movimientos"""
    df_final["Total"] = df_final.iloc[:, 11:].sum(axis=1)

    # Crear fila de totales
    fila_total = pd.DataFrame(df_final.iloc[:, 11:].sum()).T
    fila_total.insert(0, "Nro", "TOTALES")
    fila_total.insert(1, "Razon Social", "")

    return pd.concat([df_final, fila_total], ignore_index=True)


# ============================================================================
# FUNCIONES PARA GENERAR ARCHIVOS ARCA (CSV)
# ============================================================================


def obtener_conceptos_unicos(df):
    """Obtiene los conceptos únicos presentes en ventas (no NC)"""
    conceptos_en_ventas = set(
        df[df["Comprobante"] != "NC"]["Concepto"].astype(str).unique()
    )
    return sorted(list(conceptos_en_ventas))


def formatear_concepto_para_display(concepto):
    """Formatea un concepto solo para mostrar en la interfaz (quita .0 innecesarios)"""
    try:
        # Si es un número entero (ej: "106.0"), mostrar sin decimales
        if "." in str(concepto) and float(concepto) == int(float(concepto)):
            return str(int(float(concepto)))
        else:
            return str(concepto)
    except (ValueError, TypeError):
        # Si no se puede convertir a número, devolver como está
        return str(concepto)


def procesar_dataframe_para_arca(df, actividad_por_concepto):
    """Procesa el DataFrame y genera los datos para ARCA"""
    registros_salida = []

    # Conceptos que deben tener código "3"
    conceptos_3 = ["84", "85", "152"]

    # Columnas de tasas y sus respectivos códigos de alícuota
    columnas_tasas = [
        ("Tasa 21% Neto", "Tasa 21% IVA", "5"),
        ("C.F.21% Neto", "C.F.21% IVA", "5"),
        ("T.10.5% Neto", "T.10.5% IVA", "4"),
        ("C.F.10.5% Neto", "C.F.10.5% IVA", "4"),
        ("Tasa 27% Neto", "Tasa 27% IVA", "6"),
        ("R.Monot21 Neto", "R.Monot21 IVA", "5"),
        ("R.Mont.10 Neto", "R.Mont.10 IVA", "4"),
    ]

    for _, row in df.iterrows():
        # Tipo de operación según concepto
        concepto_nuevo = "2" if str(row["Concepto"]) in conceptos_3 else "1"

        # Tipo de sujeto comprador según condición
        condicion = str(row.get("Condicion", "")).strip().upper()
        if condicion == "INS.":
            tipo_sujeto = "1"
        elif condicion in ["EXE", "C.F."]:
            tipo_sujeto = "3"
        elif condicion == "MONO":
            tipo_sujeto = "2"
        else:
            tipo_sujeto = ""

        es_nc = str(row.get("Comprobante", "")).upper() == "NC"

        # Procesar tasas
        for tasa_neto_columna, tasa_iva_columna, codigo_alicuota in columnas_tasas:
            if tasa_neto_columna in df.columns and tasa_iva_columna in df.columns:
                monto_neto = row[tasa_neto_columna]
                iva = row[tasa_iva_columna]

                if monto_neto == 0 and iva == 0:
                    continue

                if es_nc:
                    monto_neto = abs(monto_neto)
                    iva = abs(iva)

                registro = {
                    "Tipo de Operacion": concepto_nuevo,
                    "Tipo de sujeto comprador": tipo_sujeto,
                    "Codigo de Alicuota": codigo_alicuota,
                    "Monto Neto Gravado": redondear_agresivo(monto_neto),
                    "Debito Fiscal Facturado": redondear_agresivo(iva),
                    "Debito Fiscal O.D.P.": redondear_agresivo(iva),
                    "Monto Neto Exento o No Gravado": 0,
                    "EsNotaCredito": es_nc,
                    "Concepto_original": str(row["Concepto"]),
                    "idx_original": row.name,
                }
                registros_salida.append(registro)

        # Exento
        if "Exento" in df.columns:
            exento = row["Exento"]
            if exento != 0:
                monto_neto_exento = abs(exento) if es_nc else exento
                registro = {
                    "Tipo de Operacion": concepto_nuevo,
                    "Tipo de sujeto comprador": tipo_sujeto,
                    "Codigo de Alicuota": "3",
                    "Monto Neto Gravado": 0,
                    "Debito Fiscal Facturado": 0,
                    "Debito Fiscal O.D.P.": 0,
                    "Monto Neto Exento o No Gravado": redondear_agresivo(
                        monto_neto_exento
                    ),
                    "EsNotaCredito": es_nc,
                    "Concepto_original": str(row["Concepto"]),
                    "idx_original": row.name,
                }
                registros_salida.append(registro)

    # Crear DataFrame completo
    df_salida = pd.DataFrame(registros_salida)

    # Eliminar filas donde 'Tipo de sujeto comprador' está vacío
    df_salida = df_salida[
        df_salida["Tipo de sujeto comprador"].astype(str).str.strip() != ""
    ]

    # Agregar la columna "Concepto" al df_salida temporalmente para poder mapear actividad
    df_salida["Concepto"] = (
        df.loc[df_salida["idx_original"], "Concepto"].astype(str).values
    )

    # Asignar columna "Actividad" según el concepto
    df_salida["Actividad"] = df_salida["Concepto"].map(actividad_por_concepto)

    # Eliminar columna "Concepto" antes de continuar
    df_salida = df_salida.drop(columns=["Concepto"])

    return df_salida


def generar_archivos_csv_arca(df_salida):
    """Genera los archivos CSV para ARCA"""
    # Separar notas de crédito y otros
    df_nc = df_salida[df_salida["EsNotaCredito"] == True].copy()
    df_otros = df_salida[df_salida["EsNotaCredito"] == False].copy()

    # Eliminar columna auxiliar antes de agrupar
    df_nc = df_nc.drop(columns=["EsNotaCredito"], errors="ignore")
    df_nc = df_nc.drop(columns=["Debito Fiscal O.D.P."], errors="ignore")
    df_otros = df_otros.drop(columns=["EsNotaCredito"], errors="ignore")

    # Agrupar incluyendo la columna "Actividad" como primer campo
    df_nc_agrupado = df_nc.groupby(
        [
            "Actividad",
            "Tipo de Operacion",
            "Tipo de sujeto comprador",
            "Codigo de Alicuota",
        ],
        as_index=False,
    ).sum()
    df_otros_agrupado = df_otros.groupby(
        [
            "Actividad",
            "Tipo de Operacion",
            "Tipo de sujeto comprador",
            "Codigo de Alicuota",
        ],
        as_index=False,
    ).sum()

    # Redondear inmediatamente después de agrupar para evitar errores de precisión acumulados
    columnas_numericas_temp = [
        "Monto Neto Gravado",
        "Debito Fiscal Facturado",
        "Debito Fiscal O.D.P.",
        "Monto Neto Exento o No Gravado",
    ]
    for col in columnas_numericas_temp:
        if col in df_nc_agrupado.columns:
            df_nc_agrupado[col] = df_nc_agrupado[col].apply(redondear_agresivo)
        if col in df_otros_agrupado.columns:
            df_otros_agrupado[col] = df_otros_agrupado[col].apply(redondear_agresivo)

    # Reordenar columnas para que "Actividad" quede primera
    column_order_nc = ["Actividad"] + [
        col for col in df_nc_agrupado.columns if col != "Actividad"
    ]
    column_order_otros = ["Actividad"] + [
        col for col in df_otros_agrupado.columns if col != "Actividad"
    ]

    df_nc_agrupado = df_nc_agrupado[column_order_nc]
    df_otros_agrupado = df_otros_agrupado[column_order_otros]

    # Limpiar columnas innecesarias
    df_nc_agrupado = df_nc_agrupado.drop(
        columns=["Concepto_original", "idx_original"], errors="ignore"
    )
    df_otros_agrupado = df_otros_agrupado.drop(
        columns=["Concepto_original", "idx_original"], errors="ignore"
    )

    # Redondear valores numéricos a 2 decimales para evitar problemas de precisión
    # Usar formateo forzado para eliminar completamente los errores de punto flotante
    columnas_numericas = [
        "Monto Neto Gravado",
        "Debito Fiscal Facturado",
        "Debito Fiscal O.D.P.",
        "Monto Neto Exento o No Gravado",
    ]

    for col in columnas_numericas:
        if col in df_nc_agrupado.columns:
            # Aplicar redondeo agresivo final para eliminar errores de precisión
            df_nc_agrupado[col] = df_nc_agrupado[col].apply(redondear_agresivo)
        if col in df_otros_agrupado.columns:
            df_otros_agrupado[col] = df_otros_agrupado[col].apply(redondear_agresivo)

    # Reemplazar ceros numéricos por cadenas vacías
    df_nc_agrupado = df_nc_agrupado.map(
        lambda x: "" if isinstance(x, (int, float)) and x == 0 else x
    )
    df_otros_agrupado = df_otros_agrupado.map(
        lambda x: "" if isinstance(x, (int, float)) and x == 0 else x
    )

    # Crear nombres únicos para los archivos
    timestamp = int(time.time())
    nombre_nc = f"archivo_rf_{timestamp}.csv"
    nombre_otros = f"archivo_df_{timestamp}.csv"

    # Exportar sin comillas en los valores
    df_nc_agrupado.to_csv(
        nombre_nc,
        index=False,
        sep=";",
        decimal=",",
        encoding="latin1",
        quoting=csv.QUOTE_NONE,
        float_format="%.2f",
    )
    df_otros_agrupado.to_csv(
        nombre_otros,
        index=False,
        sep=";",
        decimal=",",
        encoding="latin1",
        quoting=csv.QUOTE_NONE,
        float_format="%.2f",
    )

    # Función para poner comillas solo en el header
    def poner_comillas_header(nombre_archivo):
        with open(nombre_archivo, "r", encoding="latin1") as f:
            lineas = f.readlines()

        header = lineas[0].strip().split(";")
        header_comillas = ['"{}"'.format(col) for col in header]
        lineas[0] = ";".join(header_comillas) + "\n"

        with open(nombre_archivo, "w", encoding="latin1") as f:
            f.writelines(lineas)

    # Aplicar a ambos archivos
    poner_comillas_header(nombre_nc)
    poner_comillas_header(nombre_otros)

    return nombre_nc, nombre_otros, df_nc_agrupado, df_otros_agrupado


# ============================================================================
# FUNCIONES DE EXCEL
# ============================================================================


def crear_archivo_excel(df_final):
    """Crea el archivo Excel solo con la hoja de movimientos"""
    import time

    timestamp = int(time.time())
    excel_filename = f"Movimientos_{timestamp}.xlsx"

    with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
        # Solo hoja Movimientos - empezando desde la fila 1
        df_final.to_excel(writer, sheet_name="Movimientos", index=False)

    return excel_filename


def aplicar_formulas_excel(excel_filename, df_final):
    """Aplica formato de moneda en la hoja de movimientos, sin agregar ninguna fórmula de suma."""
    wb = load_workbook(excel_filename)
    wm = wb["Movimientos"]

    # La tabla ahora empieza desde la fila 1 (fila 2 considerando el encabezado)
    inicio_fila_movimientos = 2  # Fila 1 es el encabezado, datos empiezan en fila 2
    ultima_fila_movimientos = inicio_fila_movimientos + df_final.shape[0] - 1

    # Aplicar formato de moneda a todas las columnas numéricas (desde columna 11 en adelante)
    for col_idx in range(11, df_final.shape[1] + 1):
        col_letter = get_column_letter(col_idx)
        for row_idx in range(
            inicio_fila_movimientos, ultima_fila_movimientos + 1
        ):  # Formatear solo las filas de datos
            cell = wm[f"{col_letter}{row_idx}"]
            cell.number_format = '"$"#,##0.00'

    wb.save(excel_filename)


# ============================================================================
# FUNCIÓN PRINCIPAL
# ============================================================================


def procesar_archivo(file_path, tipo_esperado=None):
    """Función principal que procesa el archivo completo"""
    try:
        # 1. Leer y limpiar archivo
        lines = leer_archivo(file_path)
        encabezado_completo = procesar_encabezado(lines)
        cleaned_lines, compras_o_ventas = limpiar_lineas(lines)

        # 2. Validar que el tipo de archivo coincida con la selección
        if tipo_esperado and compras_o_ventas and compras_o_ventas != tipo_esperado:
            st.error(
                f"❌ **Error de validación**: El archivo contiene movimientos de **{compras_o_ventas}** pero seleccionaste **{tipo_esperado}**. Por favor, verifica tu selección o sube el archivo correcto."
            )
            return None, None, None

        if not compras_o_ventas:
            st.warning(
                "⚠️ No se pudo detectar automáticamente el tipo de movimientos en el archivo. Continuando con el procesamiento..."
            )
            compras_o_ventas = tipo_esperado  # Usar el tipo seleccionado por el usuario
        doble_cleaned_lines = limpiar_lineas_adicional(cleaned_lines)

        # 3. Procesar movimientos
        movements = procesar_movimientos(doble_cleaned_lines, compras_o_ventas)

        # 4. Crear DataFrames
        df = crear_dataframe_movimientos(movements)
        df_final = combinar_movimientos_duplicados(df)
        df_final = agregar_totales_movimientos(df_final)

        # Forzar tipo numérico en columnas desde la 11 en adelante
        for col in df_final.columns[11:]:
            df_final[col] = (
                pd.to_numeric(df_final[col], errors="coerce").fillna(0).astype(float)
            )

        # 5. Crear archivo Excel (sin la fila de totales)
        df_final_sin_totales = df_final[df_final["Nro"] != "TOTALES"].copy()

        # Forzar tipo float en todas las columnas numéricas posteriores a 'Jurisdiccion'
        idx_jurisdiccion = df_final_sin_totales.columns.get_loc("Jurisdiccion")
        for col in df_final_sin_totales.columns[idx_jurisdiccion + 1 :]:
            df_final_sin_totales[col] = (
                df_final_sin_totales[col]
                .astype(str)
                .str.replace(",", ".", regex=False)
                .str.replace(" ", "", regex=False)
            )
            df_final_sin_totales[col] = (
                pd.to_numeric(df_final_sin_totales[col], errors="coerce")
                .fillna(0)
                .astype(float)
            )

        excel_filename = crear_archivo_excel(df_final_sin_totales)

        # 6. Aplicar formato
        aplicar_formulas_excel(excel_filename, df_final_sin_totales)

        st.success("¡Archivo procesado con éxito!")
        return excel_filename, df_final_sin_totales, encabezado_completo

    except Exception as e:
        st.error(f"Error al procesar el archivo: {e}")
        return None, None, None


# ============================================================================
# INTERFAZ DE STREAMLIT
# ============================================================================


def main():
    st.set_page_config(
        page_title="Archivos CSV IVA Simple", page_icon="📊", layout="wide"
    )

    st.title("📊 Archivos CSV IVA Simple")
    st.markdown("---")

    # Selección del tipo de archivo
    st.subheader("🔍 Tipo de Movimientos")
    tipo_movimiento = st.selectbox(
        "Selecciona el tipo de archivo TXT que vas a procesar:",
        options=["Seleccionar...", "Compras", "Ventas"],
        help="Elige si el archivo contiene movimientos de compras o ventas",
    )

    if tipo_movimiento == "Seleccionar...":
        st.info("👆 Primero selecciona el tipo de movimientos que vas a procesar")
        st.stop()

    # Mostrar información según el tipo seleccionado
    if tipo_movimiento == "Compras":
        st.info(
            "📋 **Archivo de Compras seleccionado** - Asegúrate de subir un archivo TXT que contenga movimientos de IVA Compras"
        )
    else:
        st.info(
            "📋 **Archivo de Ventas seleccionado** - Asegúrate de subir un archivo TXT que contenga movimientos de IVA Ventas"
        )

    st.markdown("---")

    # Subir archivo TXT
    uploaded_file = st.file_uploader(
        f"Selecciona el archivo de movimientos IVA - {tipo_movimiento}",
        type=["txt"],
        help=f"Sube un archivo de texto con los movimientos IVA de {tipo_movimiento.lower()}",
    )

    if uploaded_file is None:
        st.info("👆 Sube un archivo TXT para comenzar")
        st.stop()

    # Crear ID único del archivo para cachear el procesamiento
    file_id = f"{uploaded_file.name}_{uploaded_file.size}_{tipo_movimiento}_{hash(uploaded_file.getvalue())}"

    # Solo procesar si no está en session_state
    if f"processed_{file_id}" not in st.session_state:
        # Guardar archivo temporalmente
        with open("temp_file.txt", "wb") as f:
            f.write(uploaded_file.getbuffer())

        with st.spinner("Procesando archivo..."):
            # Procesar archivo TXT
            excel_filename, df_movimientos, encabezado = procesar_archivo(
                "temp_file.txt", tipo_movimiento
            )

            # Almacenar resultados en session_state
            st.session_state[f"processed_{file_id}"] = True
            st.session_state[f"excel_{file_id}"] = excel_filename
            st.session_state[f"df_{file_id}"] = df_movimientos
            st.session_state[f"encabezado_{file_id}"] = encabezado
    else:
        # Recuperar resultados del session_state
        excel_filename = st.session_state[f"excel_{file_id}"]
        df_movimientos = st.session_state[f"df_{file_id}"]
        encabezado = st.session_state[f"encabezado_{file_id}"]

    if df_movimientos is not None:
        st.success("✅ Archivo procesado correctamente!")

        # Mostrar información del encabezado
        st.subheader("📋 Información del Archivo")
        col1, col2 = st.columns(2)

        with col1:
            st.write("**Razón Social:**", encabezado.get("RAZON SOCIAL", "N/A"))
            st.write("**CUIT:**", encabezado.get("CUIT", "N/A"))
            st.write("**Libro:**", encabezado.get("LIBRO", "N/A"))

        with col2:
            st.write("**Dirección:**", encabezado.get("DIRECCION", "N/A"))
            st.write("**Período:**", encabezado.get("PERIODO", "N/A"))
            st.write(f"**Tipo de Movimientos:** {tipo_movimiento}")

        # Mostrar estadísticas del DataFrame
        st.subheader("📊 Resumen de Datos")
        col1, col2, col3 = st.columns(3)

        with col1:
            st.metric("Total de Registros", len(df_movimientos))

        with col2:
            # Contar comprobantes únicos (excluyendo TOTALES)
            df_sin_totales = df_movimientos[df_movimientos["Nro"] != "TOTALES"]
            comprobantes_unicos = len(df_sin_totales)
            st.metric("Comprobantes Únicos", comprobantes_unicos)

        with col3:
            # Calcular total general (excluyendo fila de totales)
            if "Total" in df_movimientos.columns:
                total_general = df_movimientos[df_movimientos["Nro"] != "TOTALES"][
                    "Total"
                ].sum()
                st.metric("Total General", f"${total_general:,.2f}")

                # ========================================================================
        # SECCIÓN PARA GENERAR ARCHIVOS CSV PARA ARCA
        # ========================================================================

        st.subheader("🏛️ Generar Archivos para ARCA")

        if tipo_movimiento == "Ventas":
            st.info(
                "💡 **Información**: Esta sección te permite generar los archivos CSV necesarios para la presentación en ARCA, basados en los datos procesados del archivo TXT de **VENTAS**."
            )

            # Obtener conceptos únicos automáticamente (solo una vez por archivo)
            archivo_id = f"{tipo_movimiento}_{len(df_movimientos)}_{hash(str(df_movimientos.iloc[0].values.tobytes()) if len(df_movimientos) > 0 else 'empty')}"

            if f"conceptos_{archivo_id}" not in st.session_state:
                with st.spinner("Analizando conceptos del archivo..."):
                    conceptos_unicos = obtener_conceptos_unicos(df_movimientos)
                    st.session_state[f"conceptos_{archivo_id}"] = conceptos_unicos
            else:
                conceptos_unicos = st.session_state[f"conceptos_{archivo_id}"]

            if conceptos_unicos:
                st.success(
                    f"✅ Se encontraron {len(conceptos_unicos)} conceptos únicos"
                )

                st.write("**Asignación de códigos de actividad por concepto:**")
                st.caption(
                    "Para cada concepto encontrado en el archivo, asigna el código de actividad correspondiente."
                )

                # Crear inputs para cada concepto
                # Usar columnas para organizar mejor los inputs
                num_conceptos = len(conceptos_unicos)
                cols_per_row = 3

                st.write("Completa todos los campos antes de generar los archivos:")

                for i in range(0, num_conceptos, cols_per_row):
                    cols = st.columns(cols_per_row)
                    for j, concepto in enumerate(
                        conceptos_unicos[i : i + cols_per_row]
                    ):
                        with cols[j]:
                            # Formatear concepto solo para mostrar (quitar .0 visual)
                            concepto_display = formatear_concepto_para_display(concepto)
                            st.text_input(
                                f"Concepto {concepto_display}:",
                                key=f"concepto_{concepto}",  # Usar concepto original como key
                                help=f"Ingresa el código de actividad para el concepto {concepto_display}",
                            )

                # Botón para generar archivos CSV
                if st.button("🔄 Generar Archivos CSV para ARCA", type="primary"):
                    # Construir el diccionario solo cuando se hace click en el botón
                    actividad_por_concepto = {}
                    for concepto in conceptos_unicos:
                        codigo = st.session_state.get(f"concepto_{concepto}", "")
                        if codigo.strip():
                            actividad_por_concepto[concepto] = codigo.strip()
                    # Validar que todos los conceptos tengan código asignado
                    conceptos_sin_codigo = [
                        c
                        for c in conceptos_unicos
                        if c not in actividad_por_concepto
                        or not actividad_por_concepto[c]
                    ]

                    if conceptos_sin_codigo:
                        st.error(
                            f"❌ **Error**: Faltan códigos de actividad para los conceptos: {', '.join(conceptos_sin_codigo)}"
                        )
                    else:
                        with st.spinner("Procesando datos para ARCA..."):
                            try:
                                # Procesar DataFrame
                                df_salida = procesar_dataframe_para_arca(
                                    df_movimientos, actividad_por_concepto
                                )

                                # Generar archivos CSV
                                (
                                    nombre_nc,
                                    nombre_otros,
                                    df_nc_agrupado,
                                    df_otros_agrupado,
                                ) = generar_archivos_csv_arca(df_salida)

                                # Almacenar datos CSV en session_state para que persistan
                                csv_data_nc = None
                                csv_data_otros = None

                                if len(df_nc_agrupado) > 0:
                                    with open(nombre_nc, "rb") as f:
                                        csv_data_nc = f.read()

                                if len(df_otros_agrupado) > 0:
                                    with open(nombre_otros, "rb") as f:
                                        csv_data_otros = f.read()

                                st.session_state[f"csv_data_{file_id}"] = {
                                    "df_nc_agrupado": df_nc_agrupado,
                                    "df_otros_agrupado": df_otros_agrupado,
                                    "csv_data_nc": csv_data_nc,
                                    "csv_data_otros": csv_data_otros,
                                    "generated": True,
                                }

                                st.success("✅ ¡Archivos CSV generados correctamente!")

                                # Mostrar estadísticas
                                col1, col2 = st.columns(2)
                                with col1:
                                    st.metric(
                                        "📋 Notas de Crédito",
                                        len(df_nc_agrupado),
                                    )
                                with col2:
                                    st.metric(
                                        "📄 Otros Comprobantes",
                                        len(df_otros_agrupado),
                                    )

                            except Exception as e:
                                st.error(f"❌ Error al generar archivos CSV: {e}")

                # Mostrar botones de descarga si ya se generaron los CSV
                if f"csv_data_{file_id}" in st.session_state and st.session_state[
                    f"csv_data_{file_id}"
                ].get("generated", False):
                    csv_data = st.session_state[f"csv_data_{file_id}"]
                    df_nc_agrupado = csv_data["df_nc_agrupado"]
                    df_otros_agrupado = csv_data["df_otros_agrupado"]

                    # Botones de descarga
                    st.subheader("📥 Descargar Archivos CSV")

                    col1, col2 = st.columns(2)

                    with col1:
                        if csv_data["csv_data_nc"] is not None:
                            st.download_button(
                                label="📋 Descargar archivo_rf.csv (Notas de Crédito)",
                                data=csv_data["csv_data_nc"],
                                file_name="archivo_rf.csv",
                                mime="text/csv",
                            )
                        else:
                            st.info("No hay notas de crédito para descargar")

                    with col2:
                        if csv_data["csv_data_otros"] is not None:
                            st.download_button(
                                label="📄 Descargar archivo_df.csv (Otros Comprobantes)",
                                data=csv_data["csv_data_otros"],
                                file_name="archivo_df.csv",
                                mime="text/csv",
                            )
                        else:
                            st.info("No hay otros comprobantes para descargar")

                    # Mostrar preview de los datos
                    st.subheader("👀 Vista Previa de los Datos")

                    tab1, tab2 = st.tabs(
                        [
                            "📋 Notas de Crédito",
                            "📄 Otros Comprobantes",
                        ]
                    )

                    with tab1:
                        if len(df_nc_agrupado) > 0:
                            # Crear copia formateada para mostrar
                            df_nc_display = df_nc_agrupado.copy()
                            for col in [
                                "Monto Neto Gravado",
                                "Debito Fiscal Facturado",
                                "Debito Fiscal O.D.P.",
                                "Monto Neto Exento o No Gravado",
                            ]:
                                if col in df_nc_display.columns:
                                    df_nc_display[col] = df_nc_display[col].apply(
                                        lambda x: (
                                            f"{redondear_agresivo(x):.2f}"
                                            if pd.notnull(x) and x != 0 and x != ""
                                            else x
                                        )
                                    )
                            st.dataframe(
                                df_nc_display,
                                use_container_width=True,
                            )
                        else:
                            st.info("No hay notas de crédito en este archivo.")

                    with tab2:
                        if len(df_otros_agrupado) > 0:
                            # Crear copia formateada para mostrar
                            df_otros_display = df_otros_agrupado.copy()
                            for col in [
                                "Monto Neto Gravado",
                                "Debito Fiscal Facturado",
                                "Debito Fiscal O.D.P.",
                                "Monto Neto Exento o No Gravado",
                            ]:
                                if col in df_otros_display.columns:
                                    df_otros_display[col] = df_otros_display[col].apply(
                                        lambda x: (
                                            f"{redondear_agresivo(x):.2f}"
                                            if pd.notnull(x) and x != 0 and x != ""
                                            else x
                                        )
                                    )
                            st.dataframe(
                                df_otros_display,
                                use_container_width=True,
                            )
                        else:
                            st.info("No hay otros comprobantes en este archivo.")

            else:
                st.warning(
                    "⚠️ No se encontraron conceptos únicos en el archivo para procesar."
                )

        elif tipo_movimiento == "Compras":
            st.info(
                "💡 **Información**: Esta sección será utilizada para generar archivos CSV para **COMPRAS**."
            )
            st.warning(
                "🚧 **En Desarrollo**: La funcionalidad para procesar archivos de COMPRAS estará disponible próximamente."
            )
            st.markdown(
                """
            **Características que se implementarán:**
            
            - 📋 Procesamiento específico de conceptos de compras
            - 🔄 Generación de archivos CSV adaptados para compras
            - 📊 Validaciones específicas para movimientos de compras
            - 📥 Descarga de archivos en formato ARCA para compras
            """
            )

        else:
            st.error("❌ Tipo de movimiento no reconocido")

    else:
        st.error("❌ Error al procesar el archivo")

    # Limpiar archivos temporales

    try:
        if os.path.exists("temp_file.txt"):
            os.remove("temp_file.txt")
    except PermissionError:
        pass  # Ignorar si el archivo está siendo usado

    # Limpiar archivos Excel temporales generados
    try:
        for archivo in glob.glob("Movimientos_*.xlsx"):
            try:
                os.remove(archivo)
            except PermissionError:
                pass  # Ignorar si está siendo usado por Excel
    except Exception:
        pass

    # Limpiar archivos CSV temporales generados
    try:
        for archivo in glob.glob("archivo_rf_*.csv"):
            try:
                os.remove(archivo)
            except PermissionError:
                pass  # Ignorar si está siendo usado
        for archivo in glob.glob("archivo_df_*.csv"):
            try:
                os.remove(archivo)
            except PermissionError:
                pass  # Ignorar si está siendo usado
    except Exception:
        pass


if __name__ == "__main__":
    main()
